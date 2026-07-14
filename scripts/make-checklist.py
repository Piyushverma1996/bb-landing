from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
import os

TEAL   = "1E4A3A"
GOLD   = "C9A066"
CREAM  = "F5EDD8"
LIGHT  = "EDE0C4"
WHITE  = "FFFFFF"
RED    = "C0392B"
GREEN  = "27AE60"
ORANGE = "E67E22"
GRAY   = "F2F2F2"
DGRAY  = "CCCCCC"
BLACK  = "1A1A1A"

def teal_fill():  return PatternFill("solid", fgColor=TEAL)
def gold_fill():  return PatternFill("solid", fgColor="C9A066")
def cream_fill(): return PatternFill("solid", fgColor=CREAM)
def light_fill(): return PatternFill("solid", fgColor=LIGHT)
def gray_fill():  return PatternFill("solid", fgColor=GRAY)
def white_fill(): return PatternFill("solid", fgColor=WHITE)
def green_fill(): return PatternFill("solid", fgColor="D5F5E3")
def red_fill():   return PatternFill("solid", fgColor="FADBD8")
def orange_fill():return PatternFill("solid", fgColor="FDEBD0")

def thin_border(color=DGRAY):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom(color=TEAL):
    return Border(bottom=Side(style="medium", color=color))

def bold(size=11, color=BLACK, name="Arial"):
    return Font(bold=True, size=size, color=color, name=name)

def reg(size=10, color=BLACK, name="Arial"):
    return Font(size=size, color=color, name=name)

def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def mid():    return Alignment(horizontal="left",   vertical="center", wrap_text=False)


# ── status dropdown validation ───────────────────────────────────
def status_dv(sq):
    dv = DataValidation(
        type="list",
        formula1='"⬜ To Do,🔄 In Progress,✅ Done,⏸ Blocked"',
        allow_blank=False, showDropDown=False,
    )
    dv.sqref = sq
    return dv

# ── conditional formatting for Done rows ────────────────────────
def apply_done_cf(ws, data_start, data_end, status_col_letter, row_width_letter):
    green_rule = FormulaRule(
        formula=[f'${status_col_letter}{data_start}="✅ Done"'],
        fill=green_fill(),
    )
    ws.conditional_formatting.add(
        f"A{data_start}:{row_width_letter}{data_end}", green_rule
    )
    blocked_rule = FormulaRule(
        formula=[f'${status_col_letter}{data_start}="⏸ Blocked"'],
        fill=red_fill(),
    )
    ws.conditional_formatting.add(
        f"A{data_start}:{row_width_letter}{data_end}", blocked_rule
    )


# ════════════════════════════════════════════════════════════════
# SHEET BUILDER
# ════════════════════════════════════════════════════════════════
def write_sheet(ws, title, subtitle, col_headers, col_widths, rows, status_col_idx):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # title row
    ws.merge_cells(f"A1:{get_column_letter(len(col_headers))}1")
    ws["A1"] = f"  {title}"
    ws["A1"].font = bold(16, WHITE)
    ws["A1"].fill = teal_fill()
    ws["A1"].alignment = left()
    ws.row_dimensions[1].height = 36

    # subtitle row
    ws.merge_cells(f"A2:{get_column_letter(len(col_headers))}2")
    ws["A2"] = f"  {subtitle}"
    ws["A2"].font = reg(10, GOLD)
    ws["A2"].fill = teal_fill()
    ws["A2"].alignment = left()
    ws.row_dimensions[2].height = 20

    # blank separator
    ws.row_dimensions[3].height = 8
    for c in range(1, len(col_headers)+1):
        ws.cell(3, c).fill = gold_fill()

    # header row
    header_row = 4
    for ci, h in enumerate(col_headers, 1):
        cell = ws.cell(header_row, ci, h)
        cell.font = bold(10, WHITE)
        cell.fill = PatternFill("solid", fgColor="2A6355")
        cell.alignment = center()
        cell.border = thin_border(TEAL)
    ws.row_dimensions[header_row].height = 28

    # set column widths
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # data rows
    data_start = header_row + 1
    shade = False
    dv_cells = []

    for ri, row in enumerate(rows, data_start):
        ws.row_dimensions[ri].height = 22
        shade = ri % 2 == 0
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.font = reg(10)
            cell.alignment = left() if ci != status_col_idx else center()
            cell.border = thin_border()
            cell.fill = light_fill() if shade else white_fill()
            # colour status column
            if ci == status_col_idx:
                if val == "✅ Done":   cell.font = bold(10, GREEN)
                elif val == "⏸ Blocked": cell.font = bold(10, RED)
                elif val == "🔄 In Progress": cell.font = bold(10, ORANGE)
                else: cell.font = reg(10, "555555")
        dv_cells.append(f"{get_column_letter(status_col_idx)}{ri}")

    data_end = data_start + len(rows) - 1

    # status dropdown
    dv = status_dv(" ".join(dv_cells))
    ws.add_data_validation(dv)

    # conditional formatting
    apply_done_cf(
        ws, data_start, data_end,
        get_column_letter(status_col_idx),
        get_column_letter(len(col_headers))
    )

    return data_end


# ════════════════════════════════════════════════════════════════
# DATA
# ════════════════════════════════════════════════════════════════

# Sheet 1 — Master Setup Checklist
SETUP_HEADERS = ["#", "Category", "Task", "Who", "Effort", "Status", "Notes / Link"]
SETUP_WIDTHS  = [4, 18, 48, 14, 10, 16, 36]
SETUP_ROWS = [
    # Landing Page
    ("1",  "Landing Page", "Confirm landing page runs on localhost:3000", "Dev", "5 min",   "✅ Done",          "Run: npm run dev"),
    ("2",  "Landing Page", "Deploy to Vercel (free tier, no domain needed)", "Dev", "10 min", "🔄 In Progress",   "vercel --yes --prod"),
    ("3",  "Landing Page", "Add live Vercel URL to this sheet in Notes column (row 2)", "Both", "2 min", "⬜ To Do", "Paste URL when deployment done"),
    ("4",  "Landing Page", "Test form submission end-to-end on live URL", "Both", "5 min",  "⬜ To Do",         "Fill form → check Google Sheet"),
    ("5",  "Landing Page", "Share live URL with Urvashi for review", "Dev", "2 min",        "⬜ To Do",         "WhatsApp or email"),
    # Meta Pixel
    ("6",  "Meta Pixel",   "Go to Meta Events Manager → Create Data Source → Web Pixel", "Urvashi", "10 min", "⬜ To Do", "business.facebook.com → Events Manager"),
    ("7",  "Meta Pixel",   "Copy Pixel ID and paste into .env.local (NEXT_PUBLIC_META_PIXEL_ID)", "Dev", "5 min", "⬜ To Do", "Redeploy on Vercel after change"),
    ("8",  "Meta Pixel",   "Add Pixel ID to Vercel Dashboard → Settings → Env Variables", "Dev", "5 min", "⬜ To Do",  "Also add WEBHOOK_URL there"),
    ("9",  "Meta Pixel",   "Verify Pixel fires on landing page using Meta Pixel Helper extension", "Dev", "10 min", "⬜ To Do", "Chrome extension: Meta Pixel Helper"),
    # Make.com Automation
    ("10", "Automation",   "Create free Make.com account at make.com", "Dev", "5 min",      "⬜ To Do",         "make.com/register"),
    ("11", "Automation",   "Create new scenario: Webhook → Google Sheets → WhatsApp", "Dev", "30 min", "⬜ To Do", "See playbook Part 3 for step-by-step"),
    ("12", "Automation",   "Copy Make.com webhook URL → paste into WEBHOOK_URL env var", "Dev", "5 min", "⬜ To Do", "Vercel Dashboard + .env.local"),
    ("13", "Automation",   "Create Google Sheet 'BB Leads' with columns: Timestamp, Name, Phone, Course, Source, Status, Notes", "Urvashi", "10 min", "⬜ To Do", "Share sheet edit access with Make.com"),
    ("14", "Automation",   "Connect WhatsApp Business to Make.com (via Cloud API or WATI)", "Dev", "20 min", "⬜ To Do", "WATI 7-day free trial at wati.io"),
    ("15", "Automation",   "Test full flow: submit form → row in Sheet + WhatsApp message", "Both", "10 min", "⬜ To Do", "Use a test phone number first"),
    # Meta Ads
    ("16", "Meta Ads",     "Confirm Meta Business Manager account is active", "Urvashi", "5 min", "⬜ To Do",    "business.facebook.com"),
    ("17", "Meta Ads",     "Create Ad Account inside Business Manager if not already done", "Urvashi", "10 min", "⬜ To Do", "Add INR billing method (UPI/card)"),
    ("18", "Meta Ads",     "Create Campaign: objective = Leads, name = BB_Summer_Leads", "Dev/Urvashi", "15 min", "⬜ To Do", "Use landing page URL as destination"),
    ("19", "Meta Ads",     "Create Ad Set 1: Women 18–24, 10km Ramesh Nagar, ₹100/day", "Dev/Urvashi", "15 min", "⬜ To Do", "Interest: Beauty, Makeup, Fashion"),
    ("20", "Meta Ads",     "Create Ad Set 2: Women 25–40, 10km Ramesh Nagar, ₹100/day", "Dev/Urvashi", "15 min", "⬜ To Do", "Interest: Home beauty, Freelancing, Salon"),
    ("21", "Meta Ads",     "Upload Angle A (Student) video to Ad Set 1", "Urvashi", "10 min", "⬜ To Do",       "15-sec vertical Reel"),
    ("22", "Meta Ads",     "Upload Angle B (Homemaker) video to Ad Set 2", "Urvashi", "10 min", "⬜ To Do",     "30-sec landscape/square video"),
    ("23", "Meta Ads",     "Set campaign start date and confirm payment method", "Urvashi", "5 min", "⬜ To Do", "Start immediately once ads are approved"),
    # Instagram
    ("24", "Instagram",    "Switch Instagram to Professional (Creator/Business) account", "Urvashi", "5 min", "⬜ To Do", "Settings → Account → Switch to Professional"),
    ("25", "Instagram",    "Add landing page URL to Instagram bio", "Urvashi", "5 min",    "⬜ To Do",         "Paste Vercel URL in bio link field"),
    ("26", "Instagram",    "Connect Instagram account to Meta Business Manager", "Urvashi", "10 min", "⬜ To Do", "business.facebook.com → Accounts → Instagram"),
    ("27", "Instagram",    "Post first Reel (studio tour or course graphic)", "Urvashi", "20 min", "⬜ To Do",   "Caption template in ad brief doc"),
]

# Sheet 2 — Content Creation Tracker
CONTENT_HEADERS = ["#", "Format", "Angle / Topic", "Script Ready?", "Filmed?", "Edited?", "Posted?", "Post Date", "Platform", "Notes"]
CONTENT_WIDTHS  = [4, 14, 36, 13, 10, 10, 10, 12, 20, 28]
CONTENT_ROWS = [
    ("1",  "15s Reel",    "Angle A — Don't Waste Your Summer (Hindi)",     "✅ Done", "⬜ To Do", "⬜ To Do", "⬜ To Do", "", "Instagram Reels",         "Script in Brief doc p.3"),
    ("2",  "30s Feed Ad", "Angle B — Passion into Paycheque (Hindi)",       "✅ Done", "⬜ To Do", "⬜ To Do", "⬜ To Do", "", "Facebook + IG Feed",      "Script in Brief doc p.4"),
    ("3",  "15s Reel",    "Angle C — Career Switch (Hindi)",                 "✅ Done", "⬜ To Do", "⬜ To Do", "⬜ To Do", "", "Instagram Reels",         "Script in Brief doc p.5"),
    ("4",  "10s Story",   "Angle D — Only 2 Seats Left (film WEEKLY)",       "✅ Done", "⬜ To Do", "⬜ To Do", "⬜ To Do", "", "Stories + Reels",         "Re-film every Monday"),
    ("5",  "45s Video",   "Angle E — Founder Story (Urvashi to camera)",     "✅ Done", "⬜ To Do", "⬜ To Do", "⬜ To Do", "", "IG + FB Feed",            "Most important — retargeting"),
    ("6",  "Reel",        "Studio tour — wide shot salon pan",               "⬜ To Do","⬜ To Do", "⬜ To Do", "⬜ To Do", "", "Instagram Reels",         "Shot #2 from shot list"),
    ("7",  "Reel",        "Bridal makeup before/after reveal",                "⬜ To Do","⬜ To Do", "⬜ To Do", "⬜ To Do", "", "Instagram Reels",         "Transformation = high shares"),
    ("8",  "Carousel",    "All 3 courses — course graphic carousel",          "✅ Done","⬜ To Do",  "⬜ To Do", "⬜ To Do", "", "Instagram Feed",          "Use existing course PNGs"),
    ("9",  "Reel",        "Nail art timelapse (any design)",                  "⬜ To Do","⬜ To Do", "⬜ To Do", "⬜ To Do", "", "Instagram Reels",         "Top-down, tripod if possible"),
    ("10", "Story",       "Student testimonial (screen-record DM or video)", "⬜ To Do","⬜ To Do", "⬜ To Do", "⬜ To Do", "", "Stories + Highlights",   "Ask recent graduate"),
    ("11", "Reel",        "Urvashi teaching student 1-on-1",                  "⬜ To Do","⬜ To Do", "⬜ To Do", "⬜ To Do", "", "Instagram Reels",         "Side angle, natural convo"),
    ("12", "Post",        "Certificate ceremony — graduation moment",          "⬜ To Do","⬜ To Do", "⬜ To Do", "⬜ To Do", "", "Instagram Feed",          "Warm, celebratory caption"),
    ("13", "Reel",        "5 mistakes in bridal makeup (educational)",         "⬜ To Do","⬜ To Do", "⬜ To Do", "⬜ To Do", "", "Instagram Reels",         "Educational = saves = reach"),
    ("14", "Story",       "FAQ polls: Which course? / Budget? / Experience?", "⬜ To Do","⬜ To Do", "⬜ To Do", "⬜ To Do", "", "Instagram Stories",       "Poll stickers drive engagement"),
    ("15", "Reel",        "Wk 7 — Last 2 Seats urgency (film fresh)",         "⬜ To Do","⬜ To Do", "⬜ To Do", "⬜ To Do", "", "Reels + Stories",         "Film in week 7 specifically"),
]

# Sheet 3 — Weekly Operations
WEEKLY_HEADERS = ["Week", "Dates", "Mon", "Tue (Post)", "Wed", "Thu (Post)", "Fri", "Sat (Post)", "Sun", "Ad Check", "Done?"]
WEEKLY_WIDTHS  = [6, 16, 22, 30, 22, 30, 22, 30, 22, 34, 10]
WEEKLY_ROWS = [
    ("1", "Launch Week",   "Film all shot list", "Post: Studio Tour Reel",     "Edit videos",        "Post: Course Carousel", "Upload ads to Meta",  "Post: Meet Urvashi video",   "Review ad performance", "Launch Ad Set 1 + 2. Budget ₹200/day", "⬜ To Do"),
    ("2", "Wk 2",          "Reply all DMs",     "Post: Student testimonial",   "Film new content",   "Post: Day in the life", "Film Angle D (fresh)","Post: Nail art timelapse",   "Review leads in Sheet", "Boost top organic Reel ₹200/day",      "⬜ To Do"),
    ("3", "Wk 3",          "Reply all DMs",     "Post: Before/after bridal",   "Check ad CPL",       "Post: 5 reasons Reel",  "Kill weak ad set",    "Post: FAQ Stories (polls)",  "Count leads this week", "Kill ad set CPL >₹300. Double winner", "⬜ To Do"),
    ("4", "Wk 4",          "Reply all DMs",     "Post: Certificate ceremony",  "Prep Q&A Live",      "Post: ₹20k breakdown",  "Promote Live",        "Live Q&A 30 min on IG",      "Review pipeline",       "Launch Retargeting Ad Set (Angle E)",   "⬜ To Do"),
    ("5", "Wk 5",          "Reply all DMs",     "Post: Student transform",     "Scale winner 2x",    "Post: Nail tutorial",   "Film Angle D (fresh)","Post: Seats filling up Reel","Count total leads",     "Scale best ad set to ₹300/day",        "⬜ To Do"),
    ("6", "Wk 6",          "Reply all DMs",     "Post: New batch announce",    "Film BTS content",   "Post: BTS sourcing",    "Film Angle D (fresh)","Post: Founder Q&A Stories",  "Mid-campaign review",   "Run Angle D urgency ad separately",    "⬜ To Do"),
    ("7", "Wk 7",          "Reply all DMs",     "Post: Last 2 seats Reel",     "Prepare closing ads","Post: Graduate story",  "Max budget winner",   "Post: Summer recap carousel","Count enrolled leads",  "Max budget on best performer",         "⬜ To Do"),
    ("8", "Wk 8",          "Reply all DMs",     "Post: New batch announce",    "Prep next campaign", "Post: Early-bird offer", "Export leads list",  "Post: Thank-you video",      "Full campaign debrief", "Lookalike audience from leads list",    "⬜ To Do"),
]

# Sheet 4 — Lead Follow-Up SOP
LEAD_HEADERS = ["Step", "Trigger", "Who", "Action", "Time Limit", "Script / Template", "Status Track"]
LEAD_WIDTHS  = [6, 22, 12, 32, 13, 48, 16]
LEAD_ROWS = [
    ("1", "Form submitted",          "Make.com", "Auto-send WhatsApp confirmation to lead", "Instant",   "\"Hi [Name]! 👋 Thanks for your interest in Blushes & Brushes! Urvashi will WhatsApp you shortly to schedule your free trial 💄 📍 Ramesh Nagar, Delhi | 76784 46364\"", "Automated"),
    ("2", "Form submitted",          "Make.com", "Add row to Google Sheet BB Leads",         "Instant",   "Columns: Timestamp, Name, Phone, Course, Source, Status=New",                                     "Automated"),
    ("3", "Make.com notifies Urvashi","Urvashi", "Send personal WhatsApp voice note to lead","< 15 min",  "\"Hi [Name], main Urvashi hoon, Blushes & Brushes se. Aapka interest dekh ke bahut khushi hui! Kya hum aapki free trial class schedule kar sakte hain?\"",             "Manual"),
    ("4", "Lead responded",          "Urvashi",  "Send 2–3 student photos + salon images",   "< 30 min",  "Share: best bridal result, happy student with certificate, studio shot",                           "Manual"),
    ("5", "Lead showed interest",    "Urvashi",  "Offer 2 specific date/time slots for trial","Same day",  "\"Aapke liye Tuesday 11am ya Thursday 3pm — kaunsa better hai?\" (Always give 2 options, not open-ended)", "Manual"),
    ("6", "Trial date confirmed",    "Urvashi",  "Send confirmation + address on WhatsApp",   "Immediately","\"Confirmed! [Date] [Time] at Blushes & Brushes, B 1/1 Double Storey, Ramesh Nagar, Opp. Subway. See you then! 🎉\"", "Manual"),
    ("7", "1 day before trial",      "Urvashi",  "Send reminder WhatsApp",                   "Day before","\"Hi [Name]! Just a reminder — your free trial class is tomorrow at [Time]. Looking forward to meeting you! 💄\"", "Manual"),
    ("8", "No response after 2 days","Urvashi",  "Send single polite follow-up",             "Day 2",     "\"Hi [Name], did you get a chance to think about it? The next batch has limited seats — happy to answer any questions 😊\"", "Manual"),
    ("9", "Still no response D5",    "Urvashi",  "Final follow-up with social proof",        "Day 5",     "\"Hi [Name], sharing a quick video of one of our recent graduates [share video]. Would love to have you visit 🙏\"",  "Manual"),
    ("10","Trial attended",          "Urvashi",  "Follow up same evening for enrolment",     "Same evening","\"So lovely meeting you today! Are you ready to join the batch? I can hold your seat for 24 hours 🌸\"",             "Manual"),
    ("11","Enrolled",                "Urvashi",  "Add to WhatsApp batch group + update Sheet","Immediately","Update Google Sheet: Status → Enrolled. Add to batch WhatsApp group.",                               "Manual"),
    ("12","Not enrolling yet",       "Urvashi",  "Add to nurture list for next batch",       "Within 1 wk","Update Sheet: Status → Nurture. Follow up in 3 weeks for next batch.",                              "Manual"),
]

# Sheet 5 — Budget Tracker
BUDGET_HEADERS = ["Week", "Ad Set / Channel", "Daily Budget (₹)", "Days Running", "Spend (₹)", "Leads Generated", "Cost per Lead (₹)", "Trials Booked", "Enrolled", "Notes"]
BUDGET_WIDTHS  = [6, 26, 16, 13, 13, 16, 17, 14, 10, 30]
BUDGET_ROWS = [
    ("1", "Ad Set 1 — Students (Angle A)",     "100", "7", "=C5*D5", "0", "=IF(F5>0,E5/F5,\"—\")", "0", "0", "Launch week"),
    ("1", "Ad Set 2 — Homemakers (Angle B)",   "100", "7", "=C6*D6", "0", "=IF(F6>0,E6/F6,\"—\")", "0", "0", "Launch week"),
    ("1", "Boosted Post (top organic)",         "0",   "0", "=C7*D7", "0", "=IF(F7>0,E7/F7,\"—\")", "0", "0", "Only if strong organic post"),
    ("2", "Ad Set 1 (continued)",               "100", "7", "=C8*D8", "0", "=IF(F8>0,E8/F8,\"—\")", "0", "0", ""),
    ("2", "Ad Set 2 (continued)",               "100", "7", "=C9*D9", "0", "=IF(F9>0,E9/F9,\"—\")", "0", "0", ""),
    ("2", "Boosted Post",                       "200", "3","=C10*D10","0", "=IF(F10>0,E10/F10,\"—\")","0","0", "Boost best Reel from Wk 1"),
    ("3", "Ad Set winner (double)",             "200", "7","=C11*D11","0", "=IF(F11>0,E11/F11,\"—\")","0","0", "Kill loser ad set"),
    ("4", "Ad Set winner",                      "200", "7","=C12*D12","0", "=IF(F12>0,E12/F12,\"—\")","0","0", ""),
    ("4", "Retargeting — Angle E",              "65",  "7","=C13*D13","0", "=IF(F13>0,E13/F13,\"—\")","0","0", "Page visitors"),
    ("5", "Ad Set winner (scale 2x)",           "300", "7","=C14*D14","0", "=IF(F14>0,E14/F14,\"—\")","0","0", ""),
    ("5", "Retargeting",                        "65",  "7","=C15*D15","0", "=IF(F15>0,E15/F15,\"—\")","0","0", ""),
    ("6", "Ad Set winner",                      "300", "7","=C16*D16","0", "=IF(F16>0,E16/F16,\"—\")","0","0", ""),
    ("6", "Urgency — Angle D",                  "100", "7","=C17*D17","0", "=IF(F17>0,E17/F17,\"—\")","0","0", ""),
    ("7", "Max budget — winner",                "400", "7","=C18*D18","0", "=IF(F18>0,E18/F18,\"—\")","0","0", ""),
    ("8", "Lookalike from leads list",          "300", "7","=C19*D19","0", "=IF(F19>0,E19/F19,\"—\")","0","0", "Export leads → create lookalike"),
]


# ════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ════════════════════════════════════════════════════════════════
wb = Workbook()

# ── Sheet 1: Setup Checklist ─────────────────────────────────────
ws1 = wb.active
ws1.title = "✅ Setup Checklist"
write_sheet(ws1,
    "BLUSHES & BRUSHES — MASTER SETUP CHECKLIST",
    "Summer 2026 Lead Campaign  |  Mark tasks Done as you complete them",
    SETUP_HEADERS, SETUP_WIDTHS, SETUP_ROWS, 6)

# ── Sheet 2: Content Tracker ─────────────────────────────────────
ws2 = wb.create_sheet("🎬 Content Tracker")
write_sheet(ws2,
    "CONTENT CREATION TRACKER",
    "Track every video and post from script → filmed → edited → posted",
    CONTENT_HEADERS, CONTENT_WIDTHS, CONTENT_ROWS, 4)

# ── Sheet 3: Weekly Ops ──────────────────────────────────────────
ws3 = wb.create_sheet("📅 Weekly Ops")
write_sheet(ws3,
    "8-WEEK SUMMER OPERATIONS CALENDAR",
    "Daily task breakdown by week  |  Check off each week when complete",
    WEEKLY_HEADERS, WEEKLY_WIDTHS, WEEKLY_ROWS, 11)

# ── Sheet 4: Lead SOP ────────────────────────────────────────────
ws4 = wb.create_sheet("📲 Lead Follow-Up SOP")
write_sheet(ws4,
    "LEAD FOLLOW-UP SOP (STANDARD OPERATING PROCEDURE)",
    "Every lead that submits the form must go through all these steps in order",
    LEAD_HEADERS, LEAD_WIDTHS, LEAD_ROWS, 7)

# ── Sheet 5: Budget Tracker ──────────────────────────────────────
ws5 = wb.create_sheet("💰 Budget Tracker")

ws5.sheet_view.showGridLines = False
ws5.freeze_panes = "A4"

# header
ws5.merge_cells(f"A1:{get_column_letter(len(BUDGET_HEADERS))}1")
ws5["A1"] = "  AD SPEND & LEAD TRACKER — Summer 2026"
ws5["A1"].font = bold(16, WHITE)
ws5["A1"].fill = teal_fill()
ws5["A1"].alignment = left()
ws5.row_dimensions[1].height = 36

ws5.merge_cells(f"A2:{get_column_letter(len(BUDGET_HEADERS))}2")
ws5["A2"] = "  Fill in Leads Generated and Trials Booked each week. All other columns auto-calculate."
ws5["A2"].font = reg(10, GOLD)
ws5["A2"].fill = teal_fill()
ws5["A2"].alignment = left()
ws5.row_dimensions[2].height = 20

ws5.row_dimensions[3].height = 8
for c in range(1, len(BUDGET_HEADERS)+1):
    ws5.cell(3, c).fill = gold_fill()

for ci, h in enumerate(BUDGET_HEADERS, 1):
    cell = ws5.cell(4, ci, h)
    cell.font = bold(10, WHITE)
    cell.fill = PatternFill("solid", fgColor="2A6355")
    cell.alignment = center()
    cell.border = thin_border(TEAL)
ws5.row_dimensions[4].height = 28

for ci, w in enumerate(BUDGET_WIDTHS, 1):
    ws5.column_dimensions[get_column_letter(ci)].width = w

for ri, row in enumerate(BUDGET_ROWS, 5):
    ws5.row_dimensions[ri].height = 22
    shade = ri % 2 == 0
    for ci, val in enumerate(row, 1):
        cell = ws5.cell(ri, ci, val)
        cell.font = reg(10)
        cell.alignment = center() if ci > 2 else left()
        cell.border = thin_border()
        cell.fill = light_fill() if shade else white_fill()
        if ci == 7:  # CPL column
            cell.font = bold(10, TEAL)

# totals row
total_row = 5 + len(BUDGET_ROWS)
ws5.row_dimensions[total_row].height = 26
totals = ["TOTAL", "", "", "", f"=SUM(E5:E{total_row-1})", f"=SUM(F5:F{total_row-1})",
          f"=IF(F{total_row}>0,E{total_row}/F{total_row},\"—\")",
          f"=SUM(H5:H{total_row-1})", f"=SUM(I5:I{total_row-1})", ""]
for ci, val in enumerate(totals, 1):
    cell = ws5.cell(total_row, ci, val)
    cell.font = bold(11, WHITE)
    cell.fill = teal_fill()
    cell.alignment = center()
    cell.border = thin_border(GOLD)

# ── Dashboard Summary ────────────────────────────────────────────
ws0 = wb.create_sheet("🏠 Dashboard", 0)
ws0.sheet_view.showGridLines = False

# big title banner
ws0.merge_cells("A1:H1")
ws0["A1"] = "BLUSHES & BRUSHES"
ws0["A1"].font = Font(bold=True, size=28, color=GOLD, name="Arial")
ws0["A1"].fill = teal_fill()
ws0["A1"].alignment = center()
ws0.row_dimensions[1].height = 56

ws0.merge_cells("A2:H2")
ws0["A2"] = "Summer 2026 Lead Generation Campaign — Command Centre"
ws0["A2"].font = reg(12, WHITE)
ws0["A2"].fill = teal_fill()
ws0["A2"].alignment = center()
ws0.row_dimensions[2].height = 24

ws0.merge_cells("A3:H3")
ws0["A3"] = ""
ws0["A3"].fill = gold_fill()
ws0.row_dimensions[3].height = 6

# key metrics
ws0.row_dimensions[4].height = 14

metrics = [
    ("Total Leads",    f"='💰 Budget Tracker'!F{5+len(BUDGET_ROWS)}"),
    ("Trials Booked",  f"='💰 Budget Tracker'!H{5+len(BUDGET_ROWS)}"),
    ("Enrolled",       f"='💰 Budget Tracker'!I{5+len(BUDGET_ROWS)}"),
    ("Total Spend ₹",  f"='💰 Budget Tracker'!E{5+len(BUDGET_ROWS)}"),
    ("Avg CPL ₹",      f"='💰 Budget Tracker'!G{5+len(BUDGET_ROWS)}"),
    ("Revenue ₹",      f"='💰 Budget Tracker'!I{5+len(BUDGET_ROWS)}*20000"),
]

cols = [1, 2, 3, 5, 6, 7]
for (label, formula), col in zip(metrics, range(1, 7)):
    c = col * 1
    ws0.cell(5, c, label).font = bold(10, TEAL)
    ws0.cell(5, c, label).fill = light_fill()
    ws0.cell(5, c, label).alignment = center()
    ws0.cell(5, c, label).border = thin_border()
    ws0.cell(5, c, label).font = bold(10, "555555")

    val_cell = ws0.cell(6, c, formula)
    val_cell.font = bold(18, TEAL)
    val_cell.fill = white_fill()
    val_cell.alignment = center()
    val_cell.border = thin_border(GOLD)
    ws0.row_dimensions[6].height = 44

for c in range(1, 7):
    ws0.column_dimensions[get_column_letter(c)].width = 16

# quick links section
ws0.merge_cells("A8:F8")
ws0["A8"] = "  QUICK NAVIGATION"
ws0["A8"].font = bold(12, WHITE)
ws0["A8"].fill = PatternFill("solid", fgColor="2A6355")
ws0["A8"].alignment = left()
ws0.row_dimensions[8].height = 28

links = [
    ("✅ Setup Checklist",     "Track all technical setup tasks"),
    ("🎬 Content Tracker",    "Track every video from filmed to posted"),
    ("📅 Weekly Ops",          "Day-by-day task calendar for 8 weeks"),
    ("📲 Lead Follow-Up SOP",  "Step-by-step WhatsApp follow-up guide"),
    ("💰 Budget Tracker",      "Track ad spend, leads, and cost-per-lead"),
]
for ri, (sheet, desc) in enumerate(links, 9):
    ws0.row_dimensions[ri].height = 22
    cell = ws0.cell(ri, 1, f"→ {sheet}")
    cell.font = bold(10, TEAL)
    cell.fill = white_fill() if ri % 2 == 0 else light_fill()
    cell.alignment = left()
    cell.border = thin_border()
    desc_cell = ws0.cell(ri, 2, desc)
    desc_cell.font = reg(10, "555555")
    desc_cell.fill = white_fill() if ri % 2 == 0 else light_fill()
    desc_cell.alignment = left()
    desc_cell.border = thin_border()
    ws0.merge_cells(f"B{ri}:F{ri}")

# goal box
ws0.merge_cells("A15:F15")
ws0["A15"] = "  SUMMER CAMPAIGN GOAL"
ws0["A15"].font = bold(11, WHITE)
ws0["A15"].fill = PatternFill("solid", fgColor="2A6355")
ws0["A15"].alignment = left()
ws0.row_dimensions[15].height = 26

goals = [
    ("Budget", "₹5,000 – ₹10,000/month (₹165 – ₹330/day)"),
    ("Target Leads", "50–150 leads per month"),
    ("Target CPL", "Below ₹150 per lead"),
    ("Trial Conversion", "40–60% of leads → trial class"),
    ("Enrolment Rate", "30–50% of trial attendees → enrolled"),
    ("Campaign Period", "May – August 2026 (Summer)"),
    ("Landing Page", "Paste Vercel URL here after deployment →"),
]
for ri, (k, v) in enumerate(goals, 16):
    ws0.row_dimensions[ri].height = 20
    key = ws0.cell(ri, 1, k)
    key.font = bold(10, TEAL)
    key.fill = light_fill() if ri % 2 == 0 else white_fill()
    key.alignment = left()
    key.border = thin_border()
    val = ws0.cell(ri, 2, v)
    val.font = reg(10)
    val.fill = light_fill() if ri % 2 == 0 else white_fill()
    val.alignment = left()
    val.border = thin_border()
    ws0.merge_cells(f"B{ri}:F{ri}")

out = os.path.join(os.path.dirname(__file__), "..", "BB_Summer_Campaign_Checklist.xlsx")
wb.save(out)
print(f"Created: {out}")
